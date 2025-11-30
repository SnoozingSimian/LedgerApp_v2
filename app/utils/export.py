# app/utils/export.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime
from typing import List
from app.models.transaction import Transaction


def generate_transactions_xlsx(transactions: List[Transaction]) -> BytesIO:
    """
    Generate an XLSX file from a list of transactions.
    Returns a BytesIO object containing the Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define headers - based on actual Transaction model fields
    headers = [
        "Date",
        "Type",
        "Category",
        "Subcategory",
        "Amount",
        "Currency",
        "Payment Method",
        "Payee",
        "Merchant Type",
        "MCC Code",
        "Notes",
        "Tags",
        "Is Paid",
        "Payment Due Date",
        "Created At",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write transaction data
    for row_num, transaction in enumerate(transactions, 2):
        ws.cell(row=row_num, column=1).value = transaction.t_date.strftime("%Y-%m-%d")
        ws.cell(row=row_num, column=2).value = transaction.transaction_type.capitalize()
        ws.cell(row=row_num, column=3).value = (
            transaction.category.name if transaction.category else ""
        )
        ws.cell(row=row_num, column=4).value = (
            transaction.subcategory.name if transaction.subcategory else ""
        )
        ws.cell(row=row_num, column=5).value = float(transaction.amount)
        ws.cell(row=row_num, column=6).value = transaction.currency or "INR"
        ws.cell(row=row_num, column=7).value = transaction.payment_method or ""
        ws.cell(row=row_num, column=8).value = transaction.payee or ""
        ws.cell(row=row_num, column=9).value = transaction.merchant_type or ""
        ws.cell(row=row_num, column=10).value = transaction.mcc_code or ""
        ws.cell(row=row_num, column=11).value = transaction.notes or ""

        # Format tags as comma-separated list
        tags = (
            ", ".join([tag.tag.name for tag in transaction.tags])
            if transaction.tags
            else ""
        )
        ws.cell(row=row_num, column=12).value = tags

        ws.cell(row=row_num, column=13).value = "Yes" if transaction.is_paid else "No"
        ws.cell(row=row_num, column=14).value = (
            transaction.payment_due_date.strftime("%Y-%m-%d")
            if transaction.payment_due_date
            else ""
        )
        ws.cell(row=row_num, column=15).value = (
            transaction.created_at.strftime("%Y-%m-%d %H:%M:%S")
            if transaction.created_at
            else ""
        )

    # Adjust column widths
    column_widths = {
        "A": 12,  # Date
        "B": 10,  # Type
        "C": 20,  # Category
        "D": 20,  # Subcategory
        "E": 12,  # Amount
        "F": 10,  # Currency
        "G": 18,  # Payment Method
        "H": 25,  # Payee
        "I": 18,  # Merchant Type
        "J": 10,  # MCC Code
        "K": 35,  # Notes
        "L": 30,  # Tags
        "M": 10,  # Is Paid
        "N": 18,  # Payment Due Date
        "O": 20,  # Created At
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Format amount column as currency
    for row in range(2, len(transactions) + 2):
        ws.cell(row=row, column=5).number_format = "#,##0.00"

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file
